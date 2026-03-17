#!/usr/bin/env python3
"""
CS:GO Weapon Folder Organizer  —  parallel edition

Outputs:
  output/
  ├── inventory.json          ← raw API snapshot
  ├── database.json           ← structured tree mirroring the folder layout
  ├── database.xlsx           ← categorized spreadsheet (one sheet per category)
  ├── assets/icons/           ← .ico files only
  ├── weapons/
  │   └── <Category>/
  │       └── <Weapon>/
  │           ├── weapon_info.json  skin.png
  │           └── <Skin>/
  │               ├── skin_info.json  skin.png
  └── <InventoryCategory>/
      └── <Item>/
          ├── skin_info.json  skin.png
"""

import os, sys, json, re, struct, threading, requests
from pathlib import Path
from io import BytesIO
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import defaultdict

# ── Pillow ────────────────────────────────────────────────────────────────────
try:
    from PIL import Image
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "Pillow",
                           "--break-system-packages", "-q"])
    from PIL import Image

# ── openpyxl ──────────────────────────────────────────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl",
                           "--break-system-packages", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
BASE_WEAPONS_URL = "https://raw.githubusercontent.com/ByMykel/CSGO-API/main/public/api/en/base_weapons.json"
INVENTORY_URL    = "https://raw.githubusercontent.com/ByMykel/CSGO-API/main/public/api/en/inventory.json"
BASE_DIR         = Path("output")
WEAPONS_DIR      = BASE_DIR / "weapons"
ICONS_DIR        = BASE_DIR / "assets" / "icons"
MAX_WORKERS      = 32

CATEGORY_LABELS = {
    "skins":        None,
    "crates":       "Crates",
    "stickers":     "Stickers",
    "patches":      "Patches",
    "graffiti":     "Graffiti",
    "music_kits":   "Music Kits",
    "collectibles": "Collectibles",
    "agents":       "Agents",
    "keys":         "Keys",
    "passes":       "Passes",
    "gifts":        "Gifts",
    "tools":        "Tools",
}

# ── Thread-safe print ─────────────────────────────────────────────────────────
_print_lock = threading.Lock()
def tprint(*args, **kwargs):
    with _print_lock:
        print(*args, **kwargs)

# ── HTTP session ──────────────────────────────────────────────────────────────
_session = requests.Session()
_adapter = requests.adapters.HTTPAdapter(
    pool_connections=MAX_WORKERS, pool_maxsize=MAX_WORKERS, max_retries=3)
_session.mount("https://", _adapter)
_session.mount("http://",  _adapter)

# ── Helpers ───────────────────────────────────────────────────────────────────

def sanitize(name: str) -> str:
    return re.sub(r'[<>:"/\\|?*]', "_", name).strip()

def download_image(url: str) -> "Image.Image | None":
    try:
        r = _session.get(url, timeout=15)
        r.raise_for_status()
        return Image.open(BytesIO(r.content)).convert("RGBA")
    except Exception as e:
        tprint(f"⚠  {e}")
        return None

def save_ico(img: "Image.Image", path: Path) -> None:
    sizes = [(256,256),(128,128),(64,64),(48,48),(32,32),(16,16)]
    frames = [img.resize(s, Image.LANCZOS) for s in sizes]
    frames[0].save(path, format="ICO",
                   sizes=[(f.width, f.height) for f in frames],
                   append_images=frames[1:])

def save_png(img: "Image.Image", path: Path, size: int = 256) -> None:
    img.resize((size, size), Image.LANCZOS).save(path, format="PNG")

# ── Excel image display constants ────────────────────────────────────────────

IMG_ROW_HEIGHT = 72   # points (~96 px) — tall enough to see the image
IMG_COL_WIDTH  = 13   # chars  (~96 px)

# ── Windows folder icon ───────────────────────────────────────────────────────────

def _win_attr(path: str, attr: int) -> None:
    import ctypes
    ctypes.windll.kernel32.SetFileAttributesW(path, attr)

def set_windows_folder_icon(folder: Path, ico_path: Path) -> None:
    if os.name != "nt":
        return
    import ctypes
    ini = folder / "desktop.ini"
    if ini.exists():
        _win_attr(str(ini), 0x80)
    _win_attr(str(folder), 0x10)
    ini.write_text(
        "[.ShellClassInfo]\r\n"
        f"IconResource={ico_path.resolve()},0\r\n"
        "IconIndex=0\r\n",
        encoding="utf-8",
    )
    _win_attr(str(ini), 0x02 | 0x04)
    _win_attr(str(folder), 0x01 | 0x10)
    try:
        ctypes.windll.shell32.SHChangeNotify(0x00001000, 0x0005,
                                             str(folder.resolve()), None)
    except Exception:
        pass

# ── macOS folder icon ─────────────────────────────────────────────────────────

def _make_icns(img: "Image.Image", out_path: Path) -> None:
    buf = BytesIO()
    img.resize((256,256), Image.LANCZOS).save(buf, format="PNG")
    png = buf.getvalue()
    chunk = b"ic08" + struct.pack(">I", 8 + len(png)) + png
    out_path.write_bytes(b"icns" + struct.pack(">I", 8 + len(chunk)) + chunk)

def set_macos_folder_icon(folder: Path, img: "Image.Image") -> None:
    if sys.platform != "darwin":
        return
    import subprocess, tempfile
    with tempfile.NamedTemporaryFile(suffix=".icns", delete=False) as tmp:
        tmp_path = Path(tmp.name)
    try:
        _make_icns(img, tmp_path)
        script = (
            f'tell application "Finder"\n'
            f'  set f to POSIX file "{folder.resolve()}" as alias\n'
            f'  set src to POSIX file "{tmp_path}" as alias\n'
            f'  set the icon of f to src\n'
            f'end tell'
        )
        subprocess.run(["osascript", "-e", script], capture_output=True, timeout=10)
    except Exception:
        pass
    finally:
        tmp_path.unlink(missing_ok=True)

# ── Task ──────────────────────────────────────────────────────────────────────

class Task:
    __slots__ = ("item_folder", "ico_path", "img_url", "info", "info_filename", "label")
    def __init__(self, item_folder, ico_path, img_url, info, info_filename, label):
        self.item_folder   = item_folder
        self.ico_path      = ico_path
        self.img_url       = img_url
        self.info          = info
        self.info_filename = info_filename
        self.label         = label

def make_task(item_folder: Path, icon_key: str,
              img_url: str, info: dict, info_filename: str, label: str) -> Task:
    item_folder.mkdir(parents=True, exist_ok=True)
    ico_path = ICONS_DIR / (sanitize(icon_key) + ".ico")
    return Task(item_folder, ico_path, img_url, info, info_filename, label)

def execute_task(task: Task) -> str:
    skin_png  = task.item_folder / "skin.png"
    info_path = task.item_folder / task.info_filename
    info_path.write_text(json.dumps(task.info, indent=2, ensure_ascii=False), encoding="utf-8")

    if skin_png.exists() and task.ico_path.exists():
        return f"–  {task.label}  (skipped)"
    if not task.img_url:
        return f"✘  {task.label}  (no image URL)"

    img = download_image(task.img_url)
    if not img:
        return f"✘  {task.label}  (download failed)"

    save_ico(img, task.ico_path)
    save_png(img, skin_png)
    set_windows_folder_icon(task.item_folder, task.ico_path)
    set_macos_folder_icon(task.item_folder, img)
    return f"✔  {task.label}"

# ── Collect tasks ─────────────────────────────────────────────────────────────

def collect_weapon_tasks(weapons, skins_by_def) -> list[Task]:
    tasks = []
    categories: dict[str, list[dict]] = {}
    for w in weapons:
        cat = w.get("category", {}).get("name", "Unknown")
        categories.setdefault(cat, []).append(w)

    for cat_name, cat_weapons in sorted(categories.items()):
        cat_folder = WEAPONS_DIR / sanitize(cat_name)
        cat_folder.mkdir(parents=True, exist_ok=True)
        used_weapon_names: set[str] = set()

        for weapon in cat_weapons:
            w_name = weapon.get("name", "Unknown")
            w_id   = weapon.get("id", "unknown")
            w_def  = weapon.get("def_index")

            wfn = sanitize(w_name)
            if wfn in used_weapon_names:
                wfn = f"{wfn} ({sanitize(w_id.rsplit('-',1)[-1])})"
            used_weapon_names.add(wfn)

            w_folder = cat_folder / wfn
            w_folder.mkdir(parents=True, exist_ok=True)

            tasks.append(make_task(w_folder, w_id, weapon.get("image",""),
                                   weapon, "weapon_info.json", w_name))

            weapon_skins = skins_by_def.get(str(w_def), {})
            used_skin_names: set[str] = set()
            for paint_index, skin in weapon_skins.items():
                s_name  = skin.get("name", f"Skin {paint_index}")
                skin_fn = sanitize(s_name.split(" | ",1)[1] if " | " in s_name else s_name)
                if skin_fn in used_skin_names:
                    skin_fn = f"{skin_fn} ({paint_index})"
                used_skin_names.add(skin_fn)
                tasks.append(make_task(
                    w_folder / skin_fn,
                    f"{sanitize(w_id)}_paint{paint_index}",
                    skin.get("image",""), skin, "skin_info.json", s_name,
                ))
    return tasks

def collect_inventory_tasks(inventory) -> list[Task]:
    tasks = []
    for inv_key, inv_items in inventory.items():
        if inv_key == "skins":
            continue
        if not isinstance(inv_items, dict) or not inv_items:
            continue

        cat_label  = CATEGORY_LABELS.get(inv_key) or inv_key.replace("_"," ").title()
        cat_folder = BASE_DIR / sanitize(cat_label)
        cat_folder.mkdir(parents=True, exist_ok=True)

        # Nested = outer value is a dict AND its values are also dicts (paint_index -> item)
        # We check the first value AND confirm it has no direct item fields like "name"
        first_val = next(iter(inv_items.values()))
        is_nested = (
            isinstance(first_val, dict)
            and "name" not in first_val          # flat items always have "name"
            and any(isinstance(v, dict) for v in first_val.values())
        )
        all_items = (
            [(f"{ok}_{ik}", item)
             for ok, inner in inv_items.items() if isinstance(inner, dict)
             for ik, item in inner.items()]
            if is_nested else list(inv_items.items())
        )

        used_names: set[str] = set()
        for item_key, item in all_items:
            if not isinstance(item, dict):
                continue
            i_name = item.get("name", f"Item {item_key}")
            ifn = sanitize(i_name)
            if ifn in used_names:
                ifn = f"{ifn} ({item_key})"
            used_names.add(ifn)
            tasks.append(make_task(
                cat_folder / ifn,
                f"{inv_key}_{sanitize(item_key)}",
                item.get("image",""), item, "skin_info.json",
                f"[{cat_label}] {i_name}",
            ))
    return tasks

# ── Build structured database dict (mirrors folder tree) ─────────────────────

def build_database(weapons, skins_by_def, inventory) -> dict:
    """
    Returns a dict that mirrors the folder tree exactly:
    {
      "weapons": {
        "<Category>": {
          "<Weapon>": {
            "info": { ...weapon fields... },
            "skins": {
              "<Skin Name>": { ...skin fields... },
              ...
            }
          }
        }
      },
      "Crates": { "<Item Name>": { ...fields... } },
      "Stickers": { ... },
      ...
    }
    """
    db: dict = {"weapons": {}}

    # ── Weapons + skins ───────────────────────────────────────────────────────
    categories: dict[str, list[dict]] = {}
    for w in weapons:
        cat = w.get("category", {}).get("name", "Unknown")
        categories.setdefault(cat, []).append(w)

    for cat_name, cat_weapons in sorted(categories.items()):
        db["weapons"][cat_name] = {}
        used_weapon_names: set[str] = set()

        for weapon in cat_weapons:
            w_name = weapon.get("name", "Unknown")
            w_id   = weapon.get("id", "unknown")
            w_def  = weapon.get("def_index")

            wfn = w_name
            if wfn in used_weapon_names:
                wfn = f"{wfn} ({w_id.rsplit('-',1)[-1]})"
            used_weapon_names.add(wfn)

            weapon_skins = skins_by_def.get(str(w_def), {})
            skins_dict: dict = {}
            used_skin_names: set[str] = set()

            for paint_index, skin in weapon_skins.items():
                s_name  = skin.get("name", f"Skin {paint_index}")
                skin_fn = s_name.split(" | ",1)[1] if " | " in s_name else s_name
                if skin_fn in used_skin_names:
                    skin_fn = f"{skin_fn} ({paint_index})"
                used_skin_names.add(skin_fn)
                skins_dict[skin_fn] = skin

            db["weapons"][cat_name][wfn] = {
                "info":  weapon,
                "skins": skins_dict,
            }

    # ── Other inventory categories ─────────────────────────────────────────────
    for inv_key, inv_items in inventory.items():
        if inv_key == "skins":
            continue
        if not isinstance(inv_items, dict) or not inv_items:
            continue

        cat_label = CATEGORY_LABELS.get(inv_key) or inv_key.replace("_"," ").title()
        db[cat_label] = {}

        # Nested = outer value is a dict AND its values are also dicts (paint_index -> item)
        # We check the first value AND confirm it has no direct item fields like "name"
        first_val = next(iter(inv_items.values()))
        is_nested = (
            isinstance(first_val, dict)
            and "name" not in first_val          # flat items always have "name"
            and any(isinstance(v, dict) for v in first_val.values())
        )
        all_items = (
            [(f"{ok}_{ik}", item)
             for ok, inner in inv_items.items() if isinstance(inner, dict)
             for ik, item in inner.items()]
            if is_nested else list(inv_items.items())
        )

        used_names: set[str] = set()
        for item_key, item in all_items:
            if not isinstance(item, dict):
                continue
            i_name = item.get("name", f"Item {item_key}")
            ifn = i_name
            if ifn in used_names:
                ifn = f"{ifn} ({item_key})"
            used_names.add(ifn)
            db[cat_label][ifn] = item

    return db

# ── Excel export ──────────────────────────────────────────────────────────────

# Rarity color map (hex without #)
RARITY_COLORS = {
    "rarity_common":           "B0C3D9",
    "rarity_uncommon_weapon":  "5E98D9",
    "rarity_rare_weapon":      "4B69FF",
    "rarity_mythical_weapon":  "8847FF",
    "rarity_legendary_weapon": "D32CE6",
    "rarity_ancient_weapon":   "EB4B4B",
    "rarity_contraband":       "E4AE33",
    # non-weapon rarities
    "rarity_common_weapon":    "B0C3D9",
    "rarity_rare":             "4B69FF",
    "rarity_mythical":         "8847FF",
    "rarity_legendary":        "D32CE6",
    "rarity_ancient":          "EB4B4B",
}

def _header_style(cell, text: str, bg: str = "1B2838") -> None:
    cell.value = text
    cell.font  = Font(name="Arial", bold=True, color="C6D4DF", size=10)
    cell.fill  = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def _thin_border() -> Border:
    s = Side(style="thin", color="2A475E")
    return Border(left=s, right=s, top=s, bottom=s)

def _rarity_fill(rarity_id: str) -> PatternFill:
    color = RARITY_COLORS.get(rarity_id, "1B2838")
    return PatternFill("solid", start_color=color)

def write_weapons_sheet(ws, db_weapons: dict) -> None:
    """One sheet for all weapons + their skins, grouped by category."""
    cols = ["Image", "Category", "Weapon", "Skin Name", "Type", "Description", "Rarity", "Rarity ID", "Marketable"]
    for ci, col in enumerate(cols, 1):
        _header_style(ws.cell(1, ci), col)

    col_widths = [IMG_COL_WIDTH, 14, 22, 36, 10, 55, 16, 30, 12]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "B2"

    border = _thin_border()
    row = 2

    for cat_name, cat_weapons in db_weapons.items():
        for w_name, w_data in cat_weapons.items():
            info  = w_data["info"]
            skins = w_data["skins"]

            # Base weapon row
            img_url = info.get("image", "")
            is_weapon = True
            bg, txt_color = "16202D", "C6D4DF"
            w_desc = info.get("description", "") or ""
            row_vals = ["", cat_name, w_name, "— base —", "Weapon", w_desc, "", "", ""]
            for ci, v in enumerate(row_vals, 1):
                cell = ws.cell(row, ci, v)
                cell.font      = Font(name="Arial", size=9, color=txt_color)
                cell.fill      = PatternFill("solid", start_color=bg)
                cell.border    = border
                cell.alignment = Alignment(vertical="center", wrap_text=(ci == 6))
            if img_url:
                ws.cell(row, 1, f'=IMAGE("{img_url}")')
            ws.row_dimensions[row].height = IMG_ROW_HEIGHT
            row += 1

            # Skin rows
            for skin_name, skin in skins.items():
                rarity    = skin.get("rarity") or {}
                rarity_id = rarity.get("id", "")
                img_url   = skin.get("image", "")
                bg, txt_color = "1B2838", "FFFFFF"
                s_desc = skin.get("description", "") or ""
                row_vals = ["", cat_name, w_name, skin_name, "Skin",
                            s_desc, rarity.get("name",""), rarity_id,
                            "Yes" if skin.get("marketable") else "No"]
                for ci, v in enumerate(row_vals, 1):
                    cell = ws.cell(row, ci, v)
                    cell.font      = Font(name="Arial", size=9, color=txt_color)
                    cell.fill      = _rarity_fill(rarity_id) if ci == 7 else PatternFill("solid", start_color=bg)
                    cell.border    = border
                    cell.alignment = Alignment(vertical="center", wrap_text=(ci == 6))
                if img_url:
                    ws.cell(row, 1, f'=IMAGE("{img_url}")')
                ws.row_dimensions[row].height = IMG_ROW_HEIGHT
                row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

def write_inventory_sheet(ws, items: dict, cat_label: str) -> None:
    """One sheet per non-weapon inventory category."""
    cols = ["Image", "Name", "Description", "Rarity", "Rarity ID", "Marketable"]
    for ci, col in enumerate(cols, 1):
        _header_style(ws.cell(1, ci), col)

    col_widths = [IMG_COL_WIDTH, 36, 55, 16, 30, 12]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "B2"

    border = _thin_border()

    for row, (item_name, item) in enumerate(items.items(), 2):
        if not isinstance(item, dict):
            continue
        rarity     = item.get("rarity") or {}
        rarity_id  = rarity.get("id", "")
        img_url    = item.get("image", "") or ""
        marketable = item.get("marketable", False)
        description= item.get("description", "") or ""

        row_vals = ["", item_name, description, rarity.get("name",""), rarity_id,
                    "Yes" if marketable else "No"]
        for ci, v in enumerate(row_vals, 1):
            cell = ws.cell(row, ci, v)
            cell.font      = Font(name="Arial", size=9, color="FFFFFF")
            cell.fill      = _rarity_fill(rarity_id) if ci == 4 else PatternFill("solid", start_color="1B2838")
            cell.border    = border
            cell.alignment = Alignment(vertical="center", wrap_text=(ci == 3))
        if img_url:
            ws.cell(row, 1, f'=IMAGE("{img_url}")')
        ws.row_dimensions[row].height = IMG_ROW_HEIGHT

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

def write_summary_sheet(ws, db: dict) -> None:
    """Overview sheet: category name + item count."""
    _header_style(ws.cell(1, 1), "Category")
    _header_style(ws.cell(1, 2), "Items")
    _header_style(ws.cell(1, 3), "Skins / Sub-items")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 18
    ws.freeze_panes = "A2"
    border = _thin_border()

    row = 2
    # Weapons summary per category
    for cat_name, cat_weapons in db.get("weapons", {}).items():
        total_skins = sum(len(v["skins"]) for v in cat_weapons.values())
        for ci, v in enumerate([f"Weapons › {cat_name}", len(cat_weapons), total_skins], 1):
            cell = ws.cell(row, ci, v)
            cell.font   = Font(name="Arial", size=9, color="C6D4DF")
            cell.fill   = PatternFill("solid", start_color="16202D")
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
        row += 1

    # Other categories
    for key, items in db.items():
        if key == "weapons":
            continue
        for ci, v in enumerate([key, len(items), "—"], 1):
            cell = ws.cell(row, ci, v)
            cell.font   = Font(name="Arial", size=9, color="C6D4DF")
            cell.fill   = PatternFill("solid", start_color="1B2838")
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
        row += 1

def build_excel(db: dict, out_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Summary sheet first
    ws_summary = wb.create_sheet("Summary")
    ws_summary.sheet_properties.tabColor = "1B2838"
    write_summary_sheet(ws_summary, db)

    # Weapons sheet (all categories combined, grouped)
    ws_weapons = wb.create_sheet("Weapons")
    ws_weapons.sheet_properties.tabColor = "4B69FF"
    write_weapons_sheet(ws_weapons, db.get("weapons", {}))

    # One sheet per inventory category
    tab_colors = {
        "Crates": "E4AE33", "Stickers": "5E98D9", "Patches": "8847FF",
        "Graffiti": "EB4B4B", "Music Kits": "D32CE6", "Agents": "4B69FF",
        "Collectibles": "B0C3D9", "Keys": "C6D4DF", "Passes": "16202D",
        "Gifts": "E4AE33", "Tools": "5E98D9",
    }
    for cat_label, items in db.items():
        if cat_label == "weapons":
            continue
        ws = wb.create_sheet(cat_label[:31])  # sheet name max 31 chars
        ws.sheet_properties.tabColor = tab_colors.get(cat_label, "1B2838")
        write_inventory_sheet(ws, items, cat_label)

    wb.save(out_path)

# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    print("📡  Fetching base weapons…")
    weapons: list[dict] = _session.get(BASE_WEAPONS_URL, timeout=30).json()
    print(f"✅  {len(weapons)} weapons")

    print("📡  Fetching inventory…")
    inventory_raw = _session.get(INVENTORY_URL, timeout=60)
    inventory: dict = inventory_raw.json()
    skins_by_def    = inventory.get("skins", {})
    total_skins     = sum(len(v) for v in skins_by_def.values())
    print(f"✅  {total_skins} skins + other inventory categories\n")

    BASE_DIR.mkdir(parents=True, exist_ok=True)
    WEAPONS_DIR.mkdir(parents=True, exist_ok=True)
    ICONS_DIR.mkdir(parents=True, exist_ok=True)

    # ── Save raw inventory snapshot ───────────────────────────────────────────
    inv_path = BASE_DIR / "inventory.json"
    inv_path.write_bytes(inventory_raw.content)
    print(f"💾  inventory.json  → {inv_path}")

    # ── Build structured database ─────────────────────────────────────────────
    print("🗃   Building database…", flush=True)
    db = build_database(weapons, skins_by_def, inventory)

    db_path = BASE_DIR / "database.json"
    db_path.write_text(json.dumps(db, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"💾  database.json   → {db_path}")

    # ── Build Excel ───────────────────────────────────────────────────────────
    print("📊  Building Excel…", flush=True)
    xlsx_path = BASE_DIR / "database.xlsx"
    build_excel(db, xlsx_path)
    print(f"💾  database.xlsx   → {xlsx_path}")

    # ── Collect folder tasks (mkdir only) ─────────────────────────────────────
    print("\n🗂   Building folder tree…", flush=True)
    tasks = collect_weapon_tasks(weapons, skins_by_def)
    print(f"     weapons + skins  : {len(tasks)} items", flush=True)
    inv_tasks = collect_inventory_tasks(inventory)
    tasks += inv_tasks
    print(f"     inventory items  : {len(inv_tasks)} items", flush=True)
    print(f"\n📋  {len(tasks)} total  →  downloading with {MAX_WORKERS} threads\n", flush=True)

    # ── Parallel downloads ────────────────────────────────────────────────────
    done = 0
    total = len(tasks)
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(execute_task, t): t for t in tasks}
        for future in as_completed(futures):
            done += 1
            tprint(f"  [{done:>4}/{total}]  {future.result()}", flush=True)

    print(f"\n🎉  Done!")
    print(f"    Weapons  : {WEAPONS_DIR.resolve()}")
    print(f"    Icons    : {ICONS_DIR.resolve()}")
    print(f"    JSON DB  : {db_path.resolve()}")
    print(f"    Excel    : {xlsx_path.resolve()}")

if __name__ == "__main__":
    main()
